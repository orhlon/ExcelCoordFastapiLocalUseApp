from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.responses import FileResponse
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import tempfile
import os

app = FastAPI(docs_url=None, redoc_url=None)

@app.get("/", response_class=HTMLResponse)
async def upload_form():
    return """
    <html>
        <head>
            <title>File Upload</title>
        </head>
        <body>
            <h1>Добавьте файл и базу</h1>
            <h2>Итоговый файл с координатами появится на вашем рабочем столе.</h2>
            <h2>Программа сравнивает города в первом столбце. Координаты добавлются во 2 и 3 столбцы, вне зависимости от их содержимого.</h2>
            <h2>Если в итоговом файле присутствуют города с пустыми координатами, значит эти города необходимо добавить в базу вручную.</h2>
            <h2>Базу можно скачать по ссылке </h2>
            <a href="https://github.com/orhlon/ExcelCoordFastapiLocalUseApp/raw/refs/heads/main/BazaCoordinat.xlsx">База координат</a>
            <h2>   </h2>
            <form action="/upload/" method="post" enctype="multipart/form-data"></h2>
                <input type="file" name="file1"><br><br>
                <input type="file" name="file2"><br><br>
                <input type="submit">
            </form>
        </body>
    </html>
    """

@app.post("/upload/")
async def upload_file(file1: UploadFile = File(...), file2: UploadFile = File(...)):

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp1, \
        tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp2:
        # Save uploaded file to temp
        tmp1.write(await file1.read())
        tmp2.write(await file2.read())
        tmp1_path = tmp1.name
        tmp2_path = tmp2.name
   
    # Load workbook
    wb1 = load_workbook(tmp1_path)
    wb2 = load_workbook(tmp2_path)
    sheet1 = wb1.active
    sheet2 = wb2.active

    # Assume cities are in column A (modify as needed)
    
    for row1 in sheet1.iter_rows(min_row=2, min_col=1, max_col=1):  # Skip header
        city1 = row1[0].value
        for row2 in sheet2.iter_rows(min_row=2, min_col=1, max_col=3):
            city2 = row2[0].value
            lat = row2[1].value
            lon = row2[2].value
            if city2 == city1:
                sheet1.cell(row=row1[0].row, column=2, value=str(lat))  # Write to column B
                sheet1.cell(row=row1[0].row, column=3, value=str(lon))  # Write to column C
                print(city1 + '...' + str(lat) + '...' + str(lon))

    # Save modified workbook
    output_path = tmp1_path.replace(".xlsx", "_modified.xlsx")
    wb2.close()
    wb1.save(output_path)
    wb1.close()
    batch_command = 'copy ' + output_path + ' ' + 'C:\\users\\%username%\\desktop\\cities_with_coords.xlsx'
    print(batch_command)
    os.system(batch_command)

    # Return the modified file
    try:
        return FileResponse(
            output_path,
            filename="cities_with_coordinates.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except:
        pass


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
